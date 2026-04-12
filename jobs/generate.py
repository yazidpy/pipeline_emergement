"""
Phase 1 — GENERATE (openpyxl) : feuilles d'émargement Excel par cours.
Chemins : data/generated/YYYY-MM-DD/<session>/
Fichiers : <intitule>_<nom_prof>_<prenom_prof>_cours<id>.xlsx (suffixe _cours<id> requis pour Extract)
"""
from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import List, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Racine projet (parent de jobs/)
ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

if (ROOT / ".env").exists():
    try:
        from dotenv import load_dotenv

        load_dotenv(ROOT / ".env")
    except ImportError:
        pass

from utils.db import connect_postgres
from utils.logger import get_logger

logger = get_logger(__name__)

SESSIONS_VALIDES = ("matin", "apres-midi")

JOURS_FR: Tuple[str, ...] = (
    "Lundi",
    "Mardi",
    "Mercredi",
    "Jeudi",
    "Vendredi",
    "Samedi",
    "Dimanche",
)


def jour_semaine_fr(d: date) -> str:
    return JOURS_FR[d.weekday()]


def slug_nom_fichier(part: str, max_len: int = 80) -> str:
    """Caractères interdits Windows + espaces → underscores."""
    s = (part or "").strip()
    s = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "", s)
    s = re.sub(r"\s+", "_", s)
    s = re.sub(r"_+", "_", s).strip("_")
    if len(s) > max_len:
        s = s[:max_len].rstrip("_")
    return s or "x"


def nom_fichier_xlsx(c: CoursPlanifie) -> str:
    """
    Intitulé du cours + enseignant (nom_prénom), puis _cours<id> pour la phase Extract (PDF).
    """
    cours_p = slug_nom_fichier(c.intitule, max_len=70)
    nom_p = slug_nom_fichier(c.ens_nom, max_len=35)
    prenom_p = slug_nom_fichier(c.ens_prenom, max_len=35)
    base = f"{cours_p}_{nom_p}_{prenom_p}"
    return f"{base}_cours{c.cours_id}.xlsx"


def fmt_heure(h) -> str:
    if h is None:
        return ""
    if hasattr(h, "strftime"):
        return h.strftime("%H:%M")
    return str(h)[:5]


@dataclass(frozen=True)
class CoursPlanifie:
    cours_id: int
    intitule: str
    section: str
    horaire: str
    session: str
    heure_debut: object
    heure_fin: object
    ens_nom: str
    ens_prenom: str


@dataclass(frozen=True)
class LigneEtudiant:
    numero_et: str
    nom: str
    prenom: str
    filiere: str
    niveau: str


def fetch_cours_du_jour(session: str, jour: str) -> List[CoursPlanifie]:
    sql = """
        SELECT
            c.id,
            c.intitule,
            c.section,
            c.horaire,
            c.session,
            p.heure_debut,
            p.heure_fin,
            e.nom AS ens_nom,
            e.prenom AS ens_prenom
        FROM ref.cours c
        INNER JOIN ref.planning p ON p.cours_id = c.id
        INNER JOIN ref.enseignant e ON e.id = c.enseignant_id
        WHERE p.jour_semaine = %s AND c.session = %s
        ORDER BY c.id
    """
    out: List[CoursPlanifie] = []
    with connect_postgres() as conn:
        with conn.cursor() as cur:
            cur.execute(sql, (jour, session))
            for row in cur.fetchall():
                out.append(
                    CoursPlanifie(
                        cours_id=row[0],
                        intitule=row[1],
                        section=row[2],
                        horaire=row[3],
                        session=row[4],
                        heure_debut=row[5],
                        heure_fin=row[6],
                        ens_nom=row[7] or "",
                        ens_prenom=row[8] or "",
                    )
                )
    return out


def fetch_inscrits(cours_id: int) -> List[LigneEtudiant]:
    sql = """
        SELECT e.numero_et, e.nom, e.prenom, e.filiere, e.niveau
        FROM ref.inscription i
        INNER JOIN ref.etudiant e ON e.numero_et = i.id_etudiant
        WHERE i.cours_id = %s
        ORDER BY e.nom, e.prenom
    """
    out: List[LigneEtudiant] = []
    with connect_postgres() as conn:
        with conn.cursor() as cur:
            cur.execute(sql, (cours_id,))
            for row in cur.fetchall():
                out.append(
                    LigneEtudiant(
                        numero_et=row[0],
                        nom=row[1],
                        prenom=row[2],
                        filiere=row[3],
                        niveau=row[4],
                    )
                )
    return out


def _style_bordure_fine() -> Border:
    side = Side(style="thin", color="FF999999")
    return Border(left=side, right=side, top=side, bottom=side)


def build_workbook(
    cours: CoursPlanifie,
    date_jour: date,
    lignes: Sequence[LigneEtudiant],
) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Émargement"

    creneau = f"{fmt_heure(cours.heure_debut)}–{fmt_heure(cours.heure_fin)}"
    enseignant = f"{cours.ens_prenom} {cours.ens_nom}".strip()
    date_str = date_jour.isoformat()

    ws.merge_cells("A1:H1")
    ws["A1"] = f"Cours : {cours.intitule}"
    ws["A1"].font = Font(bold=True, size=12)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")

    ws.merge_cells("A2:H2")
    ws["A2"] = (
        f"Section : {cours.section}  |  Date : {date_str}  |  Session : {cours.session}"
    )
    ws["A3"] = f"Créneau : {creneau}  ({cours.horaire})  |  Enseignant : {enseignant}"
    ws.merge_cells("A3:H3")

    header_row = 5
    headers = (
        "Numero_et",
        "Nom",
        "Prénom",
        "Filière",
        "Niveau",
        "Présent (O/N)",
        "Signature",
        "Remarque",
    )
    fill_header = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
    font_header = Font(bold=True, color="FFFFFFFF")
    border = _style_bordure_fine()

    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.fill = fill_header
        cell.font = font_header
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    data_start = header_row + 1
    for i, etu in enumerate(lignes):
        r = data_start + i
        values = (
            etu.numero_et,
            etu.nom,
            etu.prenom,
            etu.filiere,
            etu.niveau,
            "",
            "",
            "",
        )
        for col, val in enumerate(values, start=1):
            c = ws.cell(row=r, column=col, value=val)
            c.border = border
            c.alignment = Alignment(vertical="center", wrap_text=True)

    widths = (14, 18, 14, 16, 10, 10, 22, 28)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 22
    return wb


def generer_fichiers(
    session: str,
    date_jour: date,
    dossier_base: Path | None = None,
) -> int:
    base = dossier_base or (ROOT / "data" / "generated")
    jour_label = jour_semaine_fr(date_jour)
    cours_list = fetch_cours_du_jour(session, jour_label)

    if not cours_list:
        logger.info(
            "Aucun cours pour session=%s jour=%s (%s)",
            session,
            date_jour.isoformat(),
            jour_label,
        )
        return 0

    out_dir = base / date_jour.isoformat() / session
    out_dir.mkdir(parents=True, exist_ok=True)
    n = 0

    for c in cours_list:
        inscrits = fetch_inscrits(c.cours_id)
        if not inscrits:
            logger.warning(
                "Cours id=%s sans inscrits, fichier non généré",
                c.cours_id,
            )
            continue
        wb = build_workbook(c, date_jour, inscrits)
        fname = nom_fichier_xlsx(c)
        path = out_dir / fname
        wb.save(path)
        logger.info("Généré %s (%d lignes)", path, len(inscrits))
        n += 1

    logger.info("Phase GENERATE terminée : %d fichier(s)", n)
    return n


def parse_date(s: str | None) -> date:
    if not s:
        return date.today()
    return datetime.strptime(s, "%Y-%m-%d").date()


def main(argv: List[str] | None = None) -> int:
    p = argparse.ArgumentParser(description="Génération des feuilles Excel d'émargement.")
    p.add_argument(
        "--session",
        required=True,
        choices=SESSIONS_VALIDES,
        help="matin ou apres-midi",
    )
    p.add_argument(
        "--date",
        default=None,
        help="Date au format YYYY-MM-DD (défaut : jour courant)",
    )
    args = p.parse_args(argv)

    date_jour = parse_date(args.date)
    try:
        n = generer_fichiers(args.session, date_jour)
    except Exception:
        logger.exception("Échec GENERATE")
        return 1
    return 0 if n >= 0 else 1


if __name__ == "__main__":
    sys.exit(main())
